
import openpyxl

file_path = r'C:\Users\mhass\OneDrive\Desktop\DUMP\mine\calc.xlsx'

try:
    wb = openpyxl.load_workbook(filename=file_path, data_only=False)
    sheet = wb.active
    
    print("Sheet Name:", sheet.title)
    
    # Iterate rows to find headers
    header_row = None
    gp_col_idx = None
    marks_col_idx = None
    
    for r in range(1, 20):
        row_values = [cell.value for cell in sheet[r]]
        # Check for "Grade Point" and "Obtained Marks"
        # Based on previous output, they seem to be in row 4 (1-based)
        print(f"Row {r}: {row_values}")
        
        for c_idx, val in enumerate(row_values):
            if val and "Grade Point" in str(val):
                gp_col_idx = c_idx + 1 # 1-based
                header_row = r
            if val and "Obtained Marks" in str(val):
                marks_col_idx = c_idx + 1
    
    if gp_col_idx and header_row:
        print(f"\nFound Headers at Row {header_row}. GP Col: {gp_col_idx}, Marks Col: {marks_col_idx}")
        
        # Check the next row for formula
        data_row = header_row + 1
        gp_cell = sheet.cell(row=data_row, column=gp_col_idx)
        print(f"Cell ({data_row}, {gp_col_idx}) Value: {gp_cell.value}")
        
        # If it refers to another sheet or range, we need that.
        
        # Also, scan for a lookup table on the side. 
        # Often Gpa logic is in a table like: 80 4, 79 3.9...
        # Let's read columns K onwards (11+)
        print("\n--- Lookup Table Dump (Row, Marks, GP/Other) ---")
        for r in range(1, 4):
             # Just to be sure, print header of table if possible, but let's stick to the range we suspected
             pass 

        for r in range(130, 201):
            # Print column 59 (BG), 60 (BH), 61 (BI), 62 (BJ)
            bg = sheet.cell(row=r, column=59).value
            bh = sheet.cell(row=r, column=60).value
            bi = sheet.cell(row=r, column=61).value
            bj = sheet.cell(row=r, column=62).value
            if bg or bh or bi or bj:
                print(f"Row {r}: BG={bg}, BH={bh} | BI={bi}, BJ={bj}")

except Exception as e:
    print(f"Error: {e}")
